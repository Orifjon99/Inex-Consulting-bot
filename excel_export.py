"""
Excel export utility for INEX CONSULTING Bot
Creates beautiful formatted Excel files
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)


def create_registrations_excel(registrations: List[Dict[str, Any]], filename: str = None) -> str:
    """
    Create a beautiful Excel file from registrations data

    Args:
        registrations: List of registration dictionaries
        filename: Optional filename (auto-generated if not provided)

    Returns:
        Path to the created Excel file
    """
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'inex_registrations_{timestamp}.xlsx'

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ro'yxatlar"

    # Define styles
    # Header style
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data style
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Alternating row colors
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "INEX CONSULTING - Uchrashuv Ro'yxatlari"
    title_cell.font = Font(name='Arial', size=16, bold=True, color='4472C4')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle row (date exported)
    ws.merge_cells('A2:G2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Export qilingan sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    subtitle_cell.font = Font(name='Arial', size=10, italic=True, color='666666')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ['№', 'Ism-Familiya', 'Telefon', 'Manzil', 'Korxona', 'Uchrashuv sanasi', "Ro'yxatdan o'tgan vaqt"]
    header_row = 4

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[header_row].height = 25

    # Data rows
    for row_num, registration in enumerate(registrations, start=header_row + 1):
        # Row number
        ws.cell(row=row_num, column=1).value = row_num - header_row

        # Full name
        ws.cell(row=row_num, column=2).value = registration.get('fullname', '')

        # Phone
        ws.cell(row=row_num, column=3).value = registration.get('phone', '')

        # Address
        ws.cell(row=row_num, column=4).value = registration.get('address', '')

        # Company
        ws.cell(row=row_num, column=5).value = registration.get('company', '')

        # Meeting date
        ws.cell(row=row_num, column=6).value = registration.get('meeting_date', '')

        # Created at
        ws.cell(row=row_num, column=7).value = registration.get('created_at', '')

        # Apply styling to all cells in the row
        fill = light_fill if (row_num - header_row) % 2 == 0 else white_fill

        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = fill

        # Center align row number and date columns
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row_num].height = 20

    # Column widths
    column_widths = {
        'A': 8,   # №
        'B': 25,  # Ism-Familiya
        'C': 18,  # Telefon
        'D': 30,  # Manzil
        'E': 25,  # Korxona
        'F': 20,  # Uchrashuv sanasi
        'G': 20   # Ro'yxatdan o'tgan vaqt
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Summary row
    last_row = len(registrations) + header_row + 2
    ws.merge_cells(f'A{last_row}:G{last_row}')
    summary_cell = ws.cell(row=last_row, column=1)
    summary_cell.value = f"Jami ro'yxatlar soni: {len(registrations)}"
    summary_cell.font = Font(name='Arial', size=11, bold=True, color='4472C4')
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[last_row].height = 25

    # Save file
    wb.save(filename)
    logger.info(f"Excel file created: {filename}")

    return filename
